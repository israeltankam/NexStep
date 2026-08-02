"""One hundred automated checks for the NexStep MVP."""

from __future__ import annotations

import io
import json
import os
import sys
import unittest
import uuid
import csv
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ["NEXSTEP_FAST_HASH"] = "1"

from database.connection import (
    DatabaseConfigurationError,
    adapt_query_for_postgres,
    get_connection,
    normalize_postgres_url,
)
from services.action_service import complete_action, get_next_action, list_actions, resolve_org_user_by_pin, transfer_action
from services.auth_service import identify_by_pins, password_mode, set_user_password, verify_user_password
from services.access_service import (
    authenticate_quick_access,
    create_quick_access_file,
    revoke_quick_access,
)
from services.full_backup_service import FULL_BACKUP_TABLES, export_full_database_backup
from services.comment_service import add_comment, list_comments_for_lead, search_comments
from services.lead_board_service import (
    build_lead_board,
    filter_lead_board,
    lead_board_excel,
    team_board_summary,
)
from services.lead_service import unassigned_leads_count
from services.new_lead_service import create_lead_with_first_action
from services.organization_data_service import (
    export_organization_csv_archive,
    parse_organization_csv_archive,
    replace_organization_business_data,
    verify_replacement_authorization,
)
from services.password_reset_service import (
    list_pending_requests,
    request_password_reset,
    review_password_reset,
)
from services.seed_service import AGENT_PINS, LEGACY_COMMENTS, ensure_seed_data, seed_validation_counts
from services.user_profile_service import (
    update_own_contact_details,
    update_user_contact_details_as_global_admin,
)
from utils.calendar import action_ics, google_calendar_url
from utils.dates import excel_serial_to_date, parse_date
from utils.guided_flow import action_value, due_date_from_choice, outcome_value, touchpoint_value
from utils.i18n import load_locale, t
from utils.security import hash_password, hash_pin, pin_lookup, verify_password, verify_pin
from utils.text import normalize_name, slugify
from utils.urgency import urgency_color


TABLES = [
    "organizations",
    "users",
    "organization_users",
    "pipeline_stages",
    "lead_statuses",
    "client_categories",
    "action_types",
    "leads",
    "contacts",
    "actions",
    "touchpoints",
    "transfers",
    "comments",
    "import_batches",
    "import_rows",
    "auth_attempts",
    "audit_logs",
    "auth_sessions",
    "password_reset_requests",
]

EXPECTED_COUNTS = {
    "organizations": 2,
    "agents": 5,
    "leads": 78,
    "actions": 76,
    "actions_with_due_date": 20,
    "unassigned_leads": 42,
    "legacy_comments": 22,
    "import_rows": 78,
}

CHECKS = []


def check(name):
    def decorator(func):
        CHECKS.append((name, func))
        return func

    return decorator


class NexStepCore(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp_root = ROOT / "tmp" / "tests"
        cls.tmp_root.mkdir(parents=True, exist_ok=True)
        cls.db_path = cls.tmp_root / f"nexstep_test_{uuid.uuid4().hex}.db"
        os.environ["NEXSTEP_DATABASE_PATH"] = str(cls.db_path)
        cls.conn = get_connection(cls.db_path)
        ensure_seed_data(cls.conn)
        cls.org = cls.conn.execute("SELECT * FROM organizations WHERE slug = 'les-confiotes'").fetchone()
        cls.admin_org = cls.conn.execute("SELECT * FROM organizations WHERE slug = 'scale-ag'").fetchone()

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()
        cls.db_path.unlink(missing_ok=True)
        os.environ.pop("NEXSTEP_DATABASE_PATH", None)
        os.environ.pop("NEXSTEP_FAST_HASH", None)

    def fresh_conn(self):
        path = self.tmp_root / f"fresh_{uuid.uuid4().hex}.db"
        self.addCleanup(lambda p=path: p.unlink(missing_ok=True))
        conn = get_connection(path)
        self.addCleanup(conn.close)
        ensure_seed_data(conn)
        return conn

    def agent_org_user(self, conn, display_name: str):
        return conn.execute(
            """
            SELECT ou.*, u.display_name
            FROM organization_users ou
            JOIN users u ON u.id = ou.user_id
            JOIN organizations o ON o.id = ou.organization_id
            WHERE o.slug = 'les-confiotes' AND u.display_name = ?
            """,
            (display_name,),
        ).fetchone()


for table_name in TABLES:
    @check(f"table_{table_name}_exists")
    def _(self, table_name=table_name):
        row = self.conn.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?", (table_name,)).fetchone()
        self.assertIsNotNone(row)


for key, expected in EXPECTED_COUNTS.items():
    @check(f"seed_count_{key}")
    def _(self, key=key, expected=expected):
        self.assertEqual(seed_validation_counts(self.conn)[key], expected)


for agent_name, pin in list(AGENT_PINS.items())[:1]:
    @check(f"agent_identifies_{agent_name}")
    def _(self, pin=pin):
        result = identify_by_pins(self.conn, "lesconfiotes1991", pin)
        self.assertTrue(result.ok)


@check("admin_initial_identifies")
def _(self):
    result = identify_by_pins(self.conn, "0015", "0015")
    self.assertTrue(result.ok)
    self.assertEqual(result.org_user["role"], "super_admin")


@check("bad_pin_is_generic")
def _(self):
    result = identify_by_pins(self.conn, "wrong", "wrong")
    self.assertFalse(result.ok)
    self.assertEqual(result.message_key, "login.invalid_credentials")


@check("agent_requires_password_setup")
def _(self):
    result = identify_by_pins(self.conn, "lesconfiotes1991", "0001")
    self.assertEqual(password_mode(result.user), "setup")


@check("admin_requires_initial_password_change")
def _(self):
    result = identify_by_pins(self.conn, "0015", "0015")
    self.assertEqual(password_mode(result.user), "change")


@check("set_password_allows_login")
def _(self):
    conn = self.fresh_conn()
    result = identify_by_pins(conn, "lesconfiotes1991", "0001")
    set_user_password(conn, result.user["id"], "secret")
    user = conn.execute("SELECT * FROM users WHERE id = ?", (result.user["id"],)).fetchone()
    self.assertTrue(verify_user_password(user, "secret"))


@check("pin_lookup_is_stable")
def _(self):
    self.assertEqual(pin_lookup(" 0001 "), pin_lookup("0001"))


@check("pin_hashes_are_salted")
def _(self):
    self.assertNotEqual(hash_pin("0001"), hash_pin("0001"))


@check("password_hashes_are_salted")
def _(self):
    self.assertNotEqual(hash_password("secret"), hash_password("secret"))


@check("wrong_password_fails")
def _(self):
    hashed = hash_password("secret")
    self.assertFalse(verify_password("other", hashed))


@check("pin_verification_accepts_correct_pin")
def _(self):
    hashed = hash_pin("0001")
    self.assertTrue(verify_pin("0001", hashed))


@check("normalize_duplicate_names")
def _(self):
    self.assertEqual(normalize_name("BLACK AND WHITE"), normalize_name("Black-and White"))


@check("slugify_accents")
def _(self):
    self.assertEqual(slugify("Les Confiotes!"), "les-confiotes")


@check("excel_serial_conversion")
def _(self):
    self.assertEqual(excel_serial_to_date(46204), "2026-07-01")


URGENCY_CASES = [
    ("red", "2026-07-03"),
    ("yellow", "2026-07-04"),
    ("blue", "2026-08-05"),
    ("gray", None),
]

for name, due_date in URGENCY_CASES:
    @check(f"urgency_{name}")
    def _(self, name=name, due_date=due_date):
        expected = "yellow" if name == "yellow_future" else name
        self.assertEqual(urgency_color(due_date, parse_date("2026-07-04")), expected)


@check("joel_next_action_exists")
def _(self):
    ou = self.agent_org_user(self.conn, "Joël")
    action = get_next_action(self.conn, self.org["id"], ou["id"])
    self.assertIsNotNone(action)


@check("joel_next_action_is_oldest_due")
def _(self):
    ou = self.agent_org_user(self.conn, "Joël")
    action = get_next_action(self.conn, self.org["id"], ou["id"])
    self.assertEqual(action["due_date"], "2026-05-16")


@check("list_actions_contains_urgency")
def _(self):
    ou = self.agent_org_user(self.conn, "Joël")
    actions = list_actions(self.conn, self.org["id"], org_user_id=ou["id"])
    self.assertIn("urgency_color", actions[0])


@check("list_actions_excludes_done_by_default")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    complete_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], completion_status="Oui", touchpoint_type="Appel", outcome="À relancer", note="Test", create_next=False)
    actions = list_actions(conn, self.org["id"], org_user_id=ou["id"])
    self.assertNotIn(action["id"], [item["id"] for item in actions])


@check("quick_comment_creates_row")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    add_comment(conn, organization_id=action["organization_id"], lead_id=action["lead_id"], action_id=action["id"], org_user_id=ou["id"], body="Commentaire test")
    conn.commit()
    self.assertTrue(any(c["body"] == "Commentaire test" for c in list_comments_for_lead(conn, action["lead_id"])))


@check("comment_search_finds_legacy")
def _(self):
    results = search_comments(self.conn, self.org["id"], "WhatsApp")
    self.assertGreaterEqual(len(results), 1)


@check("complete_action_marks_done")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    complete_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], completion_status="Oui", touchpoint_type="Appel", outcome="Intéressé", note="OK", create_next=False)
    row = conn.execute("SELECT status FROM actions WHERE id = ?", (action["id"],)).fetchone()
    self.assertEqual(row["status"], "done")


@check("complete_action_creates_touchpoint")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    result = complete_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], completion_status="Oui", touchpoint_type="Appel", outcome="Intéressé", note="OK", create_next=False)
    self.assertIsNotNone(result["touchpoint_id"])


@check("complete_action_creates_comment")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    complete_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], completion_status="Oui", touchpoint_type="Appel", outcome="Intéressé", note="Compte rendu test", create_next=False)
    comments = list_comments_for_lead(conn, action["lead_id"])
    self.assertTrue(any(c["comment_type"] == "action_note" for c in comments))


@check("complete_action_creates_next_action")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    result = complete_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], completion_status="Oui", touchpoint_type="Appel", outcome="À relancer", note="OK", create_next=True, next_due_date="2026-07-08", next_action_type="Appel", next_title="Relancer", next_comment="Suite")
    self.assertIsNotNone(result["next_action_id"])


@check("transfer_marks_original_transferred")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    transfer_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], target_agent_pin="0000", transfer_note="Israel connaît le contact")
    row = conn.execute("SELECT status FROM actions WHERE id = ?", (action["id"],)).fetchone()
    self.assertEqual(row["status"], "transferred")


@check("transfer_creates_target_action")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    result = transfer_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], target_agent_pin="0000", transfer_note="À reprendre")
    self.assertIsNotNone(conn.execute("SELECT id FROM actions WHERE id = ?", (result["new_action_id"],)).fetchone())


@check("transfer_comment_visible")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    action = get_next_action(conn, self.org["id"], ou["id"])
    transfer_action(conn, action_id=action["id"], actor_org_user_id=ou["id"], target_agent_pin="0000", transfer_note="À reprendre")
    comments = list_comments_for_lead(conn, action["lead_id"])
    self.assertTrue(any(c["comment_type"] == "transfer_note" for c in comments))


@check("resolve_valid_agent_pin")
def _(self):
    self.assertEqual(resolve_org_user_by_pin(self.conn, self.org["id"], "0001")["display_name"], "Joël")


@check("unassigned_leads_count")
def _(self):
    self.assertEqual(unassigned_leads_count(self.conn, self.org["id"]), 42)


@check("legacy_comments_have_source_column_a")
def _(self):
    count = self.conn.execute("SELECT COUNT(*) AS count FROM comments WHERE comment_type = 'legacy_excel_a' AND source_column = 'a'").fetchone()["count"]
    self.assertEqual(count, 22)


@check("new_lead_requires_name")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    before = conn.execute("SELECT COUNT(*) AS count FROM leads").fetchone()["count"]
    with self.assertRaises(ValueError):
        create_lead_with_first_action(
            conn,
            organization_id=self.org["id"],
            actor_org_user_id=ou["id"],
            lead_name=" ",
        )
    after = conn.execute("SELECT COUNT(*) AS count FROM leads").fetchone()["count"]
    self.assertEqual(before, after)


@check("new_lead_creates_owned_lead")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name=f"Test lead {uuid.uuid4().hex}",
        category_name="Restaurant",
    )
    lead = conn.execute("SELECT * FROM leads WHERE id = ?", (result["lead_id"],)).fetchone()
    self.assertEqual(lead["owner_org_user_id"], ou["id"])


@check("new_lead_creates_first_action")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name=f"Action lead {uuid.uuid4().hex}",
        action_type_name="Appel",
        action_title="Premier contact test",
        due_date="2026-07-06",
    )
    action = conn.execute("SELECT * FROM actions WHERE id = ?", (result["action_id"],)).fetchone()
    self.assertEqual((action["status"], action["assigned_to_org_user_id"], action["due_date"]), ("pending", ou["id"], "2026-07-06"))


@check("new_lead_creates_comment")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name=f"Comment lead {uuid.uuid4().hex}",
        context_note="Rencontré au salon.",
        action_details="Rappeler demain.",
    )
    comment = conn.execute("SELECT * FROM comments WHERE id = ?", (result["comment_id"],)).fetchone()
    self.assertIn("Rappeler demain", comment["body"])


@check("new_lead_marks_possible_duplicate")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name="Black and White",
    )
    lead = conn.execute("SELECT possible_duplicate_group FROM leads WHERE id = ?", (result["lead_id"],)).fetchone()
    self.assertEqual(lead["possible_duplicate_group"], "black-and-white")


@check("postgres_url_normalization")
def _(self):
    self.assertEqual(
        normalize_postgres_url("postgres://user:secret@example.test/db"),
        "postgresql://user:secret@example.test/db",
    )
    self.assertEqual(
        normalize_postgres_url("postgresql+psycopg://user:secret@example.test/db"),
        "postgresql://user:secret@example.test/db",
    )
    with self.assertRaises(DatabaseConfigurationError):
        normalize_postgres_url("sqlite:///local.db")


@check("guided_flow_resolves_existing_business_values")
def _(self):
    self.assertEqual(outcome_value("callback"), "À relancer")
    self.assertEqual(action_value("message"), "WhatsApp")
    self.assertIsNone(action_value("none"))
    self.assertEqual(touchpoint_value("unknown"), "Autre")


@check("guided_flow_builds_due_dates_without_database_changes")
def _(self):
    base = parse_date("2026-07-19")
    self.assertEqual(due_date_from_choice("3", base_date=base), "2026-07-22")
    self.assertIsNone(due_date_from_choice("none", base_date=base))


@check("postgres_query_adapter_preserves_literals")
def _(self):
    query = "SELECT ? AS value, '?' AS literal, \"?\" AS identifier, 'it''s ?' AS escaped"
    self.assertEqual(
        adapt_query_for_postgres(query),
        "SELECT %s AS value, '?' AS literal, \"?\" AS identifier, 'it''s ?' AS escaped",
    )


@check("cloud_requires_database_url")
def _(self):
    with mock.patch.dict(os.environ, {"APP_ENV": "cloud", "DATABASE_URL": ""}, clear=False):
        with self.assertRaises(DatabaseConfigurationError):
            get_connection()


@check("lead_board_has_one_row_per_prospect")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])
    self.assertEqual(len(rows), 78)


@check("lead_board_aggregates_contacts_actions_comments")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])
    row = next(item for item in rows if item["name"] == "BLACK AND WHITE")
    self.assertIn("contacts", row)
    self.assertIn("actions", row)
    self.assertIn("comments", row)


@check("lead_board_owner_scope")
def _(self):
    ou = self.agent_org_user(self.conn, "Joël")
    rows = build_lead_board(self.conn, self.org["id"], allowed_owner_ids=[ou["id"]])
    self.assertTrue(rows)
    self.assertEqual({row["owner_org_user_id"] for row in rows}, {ou["id"]})


@check("lead_board_searches_human_fields")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])
    filtered = filter_lead_board(rows, search="black and white")
    self.assertGreaterEqual(len(filtered), 1)


@check("lead_board_filters_urgency")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])
    filtered = filter_lead_board(rows, urgency_colors={"red"})
    self.assertTrue(filtered)
    self.assertEqual({row["urgency_color"] for row in filtered}, {"red"})


@check("lead_board_team_summary_includes_every_member")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])
    summary = team_board_summary(self.conn, self.org["id"], rows)
    self.assertEqual(len(summary), 5)
    self.assertTrue(all("lead_count" in member for member in summary))


@check("lead_board_excel_is_readable")
def _(self):
    rows = build_lead_board(self.conn, self.org["id"])[:3]
    workbook = load_workbook(io.BytesIO(lead_board_excel(rows, "fr")), read_only=True)
    self.assertEqual(workbook["Lead Board"].max_row, 4)


@check("new_lead_creates_multiple_contacts")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name=f"Multi contact {uuid.uuid4().hex}",
        contacts=[
            {"full_name": "Manager Test", "role_title": "Manager"},
            {"full_name": "Bar Test", "role_title": "Barman"},
            {"full_name": "Ops Test", "role_title": "Operations"},
        ],
    )
    self.assertEqual(len(result["contact_ids"]), 3)
    primary_count = conn.execute(
        "SELECT COUNT(*) AS count FROM contacts WHERE lead_id = ? AND is_primary = 1",
        (result["lead_id"],),
    ).fetchone()["count"]
    self.assertEqual(primary_count, 1)


@check("new_lead_preserves_contact_channels")
def _(self):
    conn = self.fresh_conn()
    ou = self.agent_org_user(conn, "Joël")
    result = create_lead_with_first_action(
        conn,
        organization_id=self.org["id"],
        actor_org_user_id=ou["id"],
        lead_name=f"Channels {uuid.uuid4().hex}",
        contacts=[
            {
                "full_name": "Awa Test",
                "phone_raw": "+237 699 000 111",
                "email": "awa@example.test",
                "whatsapp": "+237699000111",
            }
        ],
    )
    contact = conn.execute(
        "SELECT * FROM contacts WHERE id = ?",
        (result["contact_ids"][0],),
    ).fetchone()
    self.assertEqual(contact["phone_normalized"], "237699000111")
    self.assertEqual(contact["email"], "awa@example.test")
    self.assertEqual(contact["whatsapp"], "+237699000111")


@check("google_calendar_url_is_prefilled")
def _(self):
    url = google_calendar_url(title="Appeler Hôtel H", due_date="2026-07-30", details="Manager")
    self.assertIn("calendar.google.com/calendar/render", url)
    self.assertIn("20260730%2F20260731", url)
    self.assertIn("Appeler", url)


@check("ics_contains_one_day_reminder")
def _(self):
    content = action_ics(
        action_id="action-test",
        title="Rendez-vous",
        due_date="2026-07-30",
        details="Suivi",
    ).decode("utf-8")
    self.assertIn("TRIGGER:-P1D", content)
    self.assertIn("DTSTART;VALUE=DATE:20260730", content)


@check("quick_access_file_contains_no_credentials")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    _, file_bytes = create_quick_access_file(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    payload = json.loads(file_bytes)
    self.assertNotIn("pin", json.dumps(payload).casefold())
    self.assertNotIn("password", json.dumps(payload).casefold())
    stored = conn.execute("SELECT token_hash FROM auth_sessions WHERE id = ?", (payload["session_id"],)).fetchone()
    self.assertNotEqual(stored["token_hash"], payload["token"])


@check("quick_access_authenticates_active_file")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    _, file_bytes = create_quick_access_file(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    result = authenticate_quick_access(conn, file_bytes)
    self.assertTrue(result.ok)
    self.assertEqual(result.user["id"], auth.user["id"])


@check("quick_access_can_be_revoked")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    session_id, file_bytes = create_quick_access_file(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    self.assertTrue(revoke_quick_access(conn, session_id=session_id, org_user_id=auth.org_user["id"]))
    self.assertFalse(authenticate_quick_access(conn, file_bytes).ok)


@check("password_reset_request_is_deduplicated")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    conn.execute(
        "UPDATE organization_users SET role = 'company_admin' WHERE id = ?",
        (auth.org_user["id"],),
    )
    conn.commit()
    first = request_password_reset(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    second = request_password_reset(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    self.assertEqual(first, second)
    company_requests = list_pending_requests(
        conn,
        auth.organization["id"],
        reviewer_user_id=auth.user["id"],
    )
    self.assertEqual(len(company_requests), 1)
    self.assertEqual(company_requests[0]["organization_name"], "Les Confiotes")
    own_details = update_own_contact_details(
        conn,
        user_id=auth.user["id"],
        organization_id=auth.organization["id"],
        email="JOEL@example.test",
        phone="+237 600 000 001",
    )
    self.assertEqual(own_details["email"], "joel@example.test")

    global_admin = conn.execute(
        "SELECT * FROM users WHERE is_global_admin = 1"
    ).fetchone()
    global_link = conn.execute(
        "SELECT * FROM organization_users WHERE user_id = ?",
        (global_admin["id"],),
    ).fetchone()
    other_request_id = request_password_reset(
        conn,
        organization_id=global_link["organization_id"],
        user_id=global_admin["id"],
        org_user_id=global_link["id"],
    )
    company_requests = list_pending_requests(
        conn,
        auth.organization["id"],
        reviewer_user_id=auth.user["id"],
    )
    self.assertEqual(len(company_requests), 1)
    self.assertFalse(review_password_reset(
        conn,
        request_id=other_request_id,
        organization_id=auth.organization["id"],
        reviewer_user_id=auth.user["id"],
        approve=False,
    ))
    with self.assertRaises(PermissionError):
        update_user_contact_details_as_global_admin(
            conn,
            target_user_id=global_admin["id"],
            email="blocked@example.test",
            phone="",
            actor_user_id=auth.user["id"],
        )
    global_requests = list_pending_requests(
        conn,
        global_link["organization_id"],
        reviewer_user_id=global_admin["id"],
    )
    self.assertEqual(len(global_requests), 2)
    self.assertEqual(
        {row["organization_name"] for row in global_requests},
        {"Les Confiotes", "scale.ag"},
    )


@check("password_reset_approval_resets_password_and_tokens")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    set_user_password(conn, auth.user["id"], "secret")
    _, file_bytes = create_quick_access_file(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    request_id = request_password_reset(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    global_admin = conn.execute(
        "SELECT * FROM users WHERE is_global_admin = 1"
    ).fetchone()
    global_link = conn.execute(
        "SELECT * FROM organization_users WHERE user_id = ?",
        (global_admin["id"],),
    ).fetchone()
    self.assertTrue(review_password_reset(
        conn,
        request_id=request_id,
        organization_id=global_link["organization_id"],
        reviewer_user_id=global_admin["id"],
        approve=True,
    ))
    user = conn.execute("SELECT * FROM users WHERE id = ?", (auth.user["id"],)).fetchone()
    self.assertEqual(password_mode(user), "setup")
    self.assertFalse(authenticate_quick_access(conn, file_bytes).ok)
    audit = conn.execute(
        "SELECT * FROM audit_logs WHERE entity_id = ?",
        (request_id,),
    ).fetchone()
    self.assertEqual(audit["organization_id"], auth.organization["id"])
    self.assertEqual(audit["actor_user_id"], global_admin["id"])
    updated_details = update_user_contact_details_as_global_admin(
        conn,
        target_user_id=auth.user["id"],
        email="joel.updated@example.test",
        phone="+237 600 000 002",
        actor_user_id=global_admin["id"],
    )
    self.assertEqual(updated_details["phone"], "+237 600 000 002")


@check("company_csv_archive_round_trip")
def _(self):
    archive = export_organization_csv_archive(self.conn, self.org["id"])
    parsed = parse_organization_csv_archive(archive, self.org["id"])
    self.assertEqual(len(parsed["leads"]), 78)
    self.assertEqual(len(parsed["comments"]), 22)


@check("company_csv_archive_rejects_other_company")
def _(self):
    archive = export_organization_csv_archive(self.conn, self.org["id"])
    with self.assertRaisesRegex(ValueError, "wrong_organization"):
        parse_organization_csv_archive(archive, self.admin_org["id"])


@check("company_replacement_requires_three_pins_and_password")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    set_user_password(conn, auth.user["id"], "admin-secret")
    self.assertFalse(
        verify_replacement_authorization(
            conn,
            organization_id=auth.organization["id"],
            user_id=auth.user["id"],
            company_pins=("lesconfiotes1991", "wrong", "lesconfiotes1991"),
            password="admin-secret",
        )
    )
    self.assertTrue(
        verify_replacement_authorization(
            conn,
            organization_id=auth.organization["id"],
            user_id=auth.user["id"],
            company_pins=("lesconfiotes1991",) * 3,
            password="admin-secret",
        )
    )


@check("company_replacement_preserves_accounts_and_sessions")
def _(self):
    conn = self.fresh_conn()
    auth = identify_by_pins(conn, "lesconfiotes1991", "0001")
    create_quick_access_file(
        conn,
        organization_id=auth.organization["id"],
        user_id=auth.user["id"],
        org_user_id=auth.org_user["id"],
    )
    archive = export_organization_csv_archive(conn, auth.organization["id"])
    parsed = parse_organization_csv_archive(archive, auth.organization["id"])
    lead = conn.execute(
        "SELECT id, name FROM leads WHERE organization_id = ? ORDER BY id LIMIT 1",
        (auth.organization["id"],),
    ).fetchone()
    users_before = conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"]
    conn.execute("UPDATE leads SET name = 'Temporary change' WHERE id = ?", (lead["id"],))
    conn.commit()
    replace_organization_business_data(
        conn,
        organization_id=auth.organization["id"],
        actor_user_id=auth.user["id"],
        archive_data=parsed,
    )
    restored = conn.execute("SELECT name FROM leads WHERE id = ?", (lead["id"],)).fetchone()
    users_after = conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()["count"]
    sessions = conn.execute("SELECT COUNT(*) AS count FROM auth_sessions").fetchone()["count"]
    self.assertEqual(restored["name"], lead["name"])
    self.assertEqual(users_after, users_before)
    self.assertEqual(sessions, 1)


@check("full_backup_contains_every_nexstep_table")
def _(self):
    archive_bytes = export_full_database_backup(self.conn)
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
        names = set(archive.namelist())
        expected = {f"tables/{table}.csv" for table in FULL_BACKUP_TABLES}
        self.assertTrue(expected.issubset(names))
        manifest = list(
            csv.DictReader(
                io.StringIO(archive.read("manifest.csv").decode("utf-8-sig"))
            )
        )
    self.assertEqual({row["table_name"] for row in manifest}, set(FULL_BACKUP_TABLES))


@check("full_backup_covers_all_organizations_and_users")
def _(self):
    archive_bytes = export_full_database_backup(self.conn)
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
        organizations = list(
            csv.DictReader(
                io.StringIO(archive.read("tables/organizations.csv").decode("utf-8-sig"))
            )
        )
        users = list(
            csv.DictReader(
                io.StringIO(archive.read("tables/users.csv").decode("utf-8-sig"))
            )
        )
        self.assertIn("schema.sql", archive.namelist())
        self.assertIn("IMPORTANT_README.txt", archive.namelist())
    self.assertEqual(len(organizations), 2)
    self.assertEqual(len(users), 6)


@check("supabase_security_covers_all_tables")
def _(self):
    security_sql = (ROOT / "database" / "supabase_security.sql").read_text(encoding="utf-8")
    for table in TABLES:
        self.assertIn(f"ALTER TABLE public.{table} ENABLE ROW LEVEL SECURITY", security_sql)
    self.assertIn("FROM anon, authenticated", security_sql)


I18N_KEYS = [
    "login.company_pin",
    "nav.new_lead",
    "guided.outcome_question",
]

for language in ("fr", "en"):
    for key in I18N_KEYS:
        @check(f"i18n_{language}_{key}")
        def _(self, language=language, key=key):
            self.assertNotEqual(t(key, language), key)


def _make_test(func):
    def test(self):
        return func(self)

    return test


assert len(CHECKS) == 100, f"Expected exactly 100 generated checks, got {len(CHECKS)}"

for index, (name, func) in enumerate(CHECKS, start=1):
    setattr(NexStepCore, f"test_{index:03d}_{name}", _make_test(func))


if __name__ == "__main__":
    unittest.main(verbosity=2)
